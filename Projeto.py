from sklearn.datasets import load_breast_cancer
from sklearn.naive_bayes import GaussianNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import cross_validate, KFold
from numpy import mean
from pandas import DataFrame
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np 
import pylab 
import scipy.stats as st
import openpyxl

#Carrega o Data Set
cancer = load_breast_cancer()

# Cria o dataframe
df_cancer = pd.DataFrame(data=cancer.data, columns=cancer.feature_names)
df_cancer['class'] = cancer.target

#Divide a coluna da classe das demais
Y = df_cancer.pop('class')
X = df_cancer
#k para a validação cruzada
k=5

# Outro data set
names = ['ID','Clump Thickness','Uniformity of Cell Size','Uniformity of Cell Shape','Marginal Adhesion','Single Epithelial Cell Size','Bare Nuclei', 'Bland Chromatin','Normal Nucleoli','Mitoses','Class']
cancer = pd.read_csv("breast-cancer-wisconsin.data",names = names, header=None)
cancer.pop('ID')
y = cancer.pop('Class')
x = cancer
dataset = 0

wb = openpyxl.Workbook()
wb.save('resultados.xlsx')
for _ in range(2):
    k=5
    for _ in range(2):
        #Lista dos resultados
        Lista_NB_acuracia = []
        Lista_NB_recall = []
        Lista_NB_tempo = []
        Lista_KNN_acuracia = []
        Lista_KNN_recall = []
        Lista_KNN_tempo = []
        for _ in range(30):
            #Naive Bayes utilizando validação cruzada
            modelo=GaussianNB()
            cv = KFold(n_splits=k, shuffle=True)
            scoring=['accuracy', 'recall']
            if dataset==0:
                scores = cross_validate(modelo, X,Y, scoring=scoring, cv=cv)
            else:
                scores = cross_validate(modelo, x,y, scoring=scoring, cv=cv)

            # Coloca na lista os resultados obtidos
            Lista_NB_acuracia.append(mean(scores['test_accuracy'])*100)
            Lista_NB_recall.append(mean(scores['test_recall'])*100)
            Lista_NB_tempo.append(sum(scores['fit_time']+scores['score_time']))
            
            # KNN com validação cruzada
            modelo = KNeighborsClassifier(n_neighbors=5, metric='euclidean')
            cv = KFold(n_splits=k, shuffle=True)
            scoring=['accuracy','recall']
            if dataset==0:
                scores = cross_validate(modelo, X,Y, scoring=scoring, cv=cv)
            else:
                scores = cross_validate(modelo, x,y, scoring=scoring, cv=cv)

            # Coloca na lista os resultados obtidos
            Lista_KNN_acuracia.append(mean(scores['test_accuracy'])*100)
            Lista_KNN_recall.append(mean(scores['test_recall'])*100)
            Lista_KNN_tempo.append(sum(scores['fit_time']+scores['score_time']))

        
        #cria um dataframe com os resultados e depois passa para uma tabela
        wb=openpyxl.load_workbook("resultados.xlsx")
        book = load_workbook("resultados.xlsx")
        writer = pd.ExcelWriter("resultados.xlsx", engine = 'openpyxl')
        writer.book = book
        #writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        df = DataFrame({'Amostras': range(1,31),'Acurácia Naive Bayes': Lista_NB_acuracia, 'Recall Naive Bayes': Lista_NB_recall,'Tempo Naive Bayes':Lista_NB_tempo,'Acurácia Knn': Lista_KNN_acuracia, 'Recall Knn': Lista_KNN_recall,'Tempo Knn':Lista_KNN_tempo})
        if k==5:
            if dataset==0:
                df.to_excel(writer, sheet_name='k=5 e dataset = 0', index=False)
            else:
                df.to_excel(writer, sheet_name='k=5 e dataset = 1', index=False)
        else:
            if dataset==0:
                df.to_excel(writer, sheet_name='k=10 e dataset = 0', index=False)
            else:
                df.to_excel(writer, sheet_name='k=10 e dataset = 1', index=False)
        writer.save()
        writer.close()
        # Lista com as médias e desvio padrão de cada métrica
        medias=[mean(Lista_NB_acuracia),mean(Lista_NB_recall),mean(Lista_NB_tempo),mean(Lista_KNN_acuracia),mean(Lista_KNN_recall),mean(Lista_KNN_tempo)]
        desvios=[np.std(Lista_NB_acuracia),np.std(Lista_NB_recall),np.std(Lista_NB_tempo),np.std(Lista_KNN_acuracia),np.std(Lista_KNN_recall),np.std(Lista_KNN_tempo)]
        # Lista dos resultado
        Lista_resultados=[Lista_NB_acuracia,Lista_NB_recall,Lista_NB_tempo,Lista_KNN_acuracia, Lista_KNN_recall,Lista_KNN_tempo]
        # Lista do nomes das métricas
        nome=['Acurácia NB','Recall NB','Tempo NB','Acurácia Knn','Recall Knn','Tempo Knn']

        #Para cada métrica
        for j in range(6):
            # lista para os decis
            decis_amostrais=[] 
            decis_teoricos=[]
            teorico_normal=[]
            i=0.1
            # para cada decil(0,1 -0,9)
            for _ in range(9):
                #calcula o quantil amostral(Yi)
                decis_amostrais.append(np.quantile(Lista_resultados[j],i))
                #calcula o quantil teorico para normal
                teorico_normal.append(st.norm.ppf(i))
                #calcula o quantil teorico amostral(Xi)
                decis_teoricos.append((desvios[j]*st.norm.ppf(i))+medias[j])
                i+=0.1

            #Criar o gráfico
            graf = pd.DataFrame({'Decis amostrais' : decis_amostrais,'Decis teóricos' : decis_teoricos})
            graf[['Decis amostrais']].plot(color='blue',kind='bar', width = 0.4)
            graf['Decis teóricos'].plot(color='orange',kind='line',secondary_y=False,label='Decis téoricos')
            ax = plt.gca()
            plt.legend()
            plt.xlim([-0.4, 9-0.4])
            if j!=2 and j!=5:
                plt.ylim([0,200])
            if k == 5:
                if dataset==0:
                    plt.title('Q-Q Plot Normal para '+nome[j]+' K=5 e DataSet0')
                else:
                    plt.title('Q-Q Plot Normal para '+nome[j]+' K=5 e DataSet1')
            else:
                if dataset==0:
                    plt.title('Q-Q Plot Normal para '+nome[j]+' K=10 e DataSet0')
                else:
                    plt.title('Q-Q Plot Normal para '+nome[j]+' K=10 e DataSet1')
            
            ax.set_xticklabels(('0.1','0.2','0.3','0.4','0.5','0.6','0.7','0.8','0.9'))
            for index,data in enumerate(decis_amostrais):
                plt.text(x=index , y =data-(data*.005) , s=round(data,2) , fontdict=dict(fontsize=7))
            for index,data in enumerate(decis_teoricos):
                plt.text(x=index , y =data+(data*.005), s=round(data,2) , fontdict=dict(fontsize=7))
            #plt.show()
            if k == 5:
                if dataset==0:
                    plt.savefig('Gráfico Quantil-Quantil Normal '+nome[j]+' K=5 e DataSet0'+'.png')
                else:
                    plt.savefig('Gráfico Quantil-Quantil Normal '+nome[j]+' K=5 e DataSet1'+'.png')
            else:
                if dataset==0:
                    plt.savefig('Gráfico Quantil-Quantil Normal '+nome[j]+' K=10 e DataSet0'+'.png')
                else:
                    plt.savefig('Gráfico Quantil-Quantil Normal '+nome[j]+' K=10 e DataSet1'+'.png')
            plt.close()
            
        k=10
    dataset=1

        

